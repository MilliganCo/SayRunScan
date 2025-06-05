# barcode_backend.py – чтение .xlsx в память + кэш заказов + название номенклатуры
# --------------------------------------------------------------
#  ▸ кэшируем:
#       – wh[barcode]  → {stock:int, name:str}
#       – wb[barcode]  → qty (WB остатки)
#       – ozon[barcode]→ qty (Ozon остатки)
#       – wb_yes / wb_week / ozon_week – статистика заказов
#  ▸ «неделя» = прошлый календарный промежуток Пн‑Вс (7 дней)
#  ▸ обновляем всё раз в час или по POST /refresh

import os, glob, io, csv, json, requests, psycopg2
from datetime import datetime
from flask import Flask, request, jsonify, abort
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import load_workbook

# ------------------- конфигурация -------------------
DB_CFG_PATH   = r"\\194.32.248.34\Shared\Keys\db.json"
WB_JSON_PATH  = r"\\194.32.248.34\Shared\wb_stocks_new.json"
OZON_META_PATH= r"\\194.32.248.34\Shared\ozon_stocks_path.json"
ONE_C_DIR     = r"\\194.32.248.34\Shared\1C"
AUTH_TOKEN    = "HvkhvUVUhvuvuYVUKvukyV"



with open(DB_CFG_PATH, encoding="utf-8") as f:
    DB = json.load(f)

app = Flask(__name__)

# ---------- глобальные кэши ----------
wh_cache:   dict[str, dict] = {}
wb_cache:   dict[str, int]  = {}
ozon_cache: dict[str, int]  = {}
wb_yes_cache   : dict[str,int] = {}
wb_week_cache  : dict[str,int] = {}
ozon_week_cache: dict[str,int] = {}
wb_supply_cache: dict[str,int] = {}  # Новый кэш для поставок
last_refresh: datetime | None = None
# -------------------------------------

def db_conn():
    return psycopg2.connect(**DB)

# ------------------- loaders -------------------

def _load_wh_cache() -> dict[str, dict]:
    """Считываем новейший .xlsx из 1С прямо в память, колонка A – name, G – barcode, I – stock"""
    newest = max(glob.glob(os.path.join(ONE_C_DIR, "*.xlsx")), key=os.path.getmtime, default=None)
    if not newest:
        return {}
    with open(newest, "rb") as f:
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
    ws = wb.active
    data: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        bc   = str(row[6]).split(".")[0] if row[6] else ""
        stock= int(row[8] or 0)
        if bc:
            data[bc] = {"stock": stock, "name": name}
    wb.close()
    return data


def _load_wb_cache() -> dict[str,int]:
    data: dict[str,int] = {}
    with open(WB_JSON_PATH, encoding="utf-8") as f:
        arr = json.load(f)
    for item in arr:
        bc = str(item.get("barcode", "")).strip()
        for w in item.get("warehouses", []):
            if w.get("warehouseName") == "Всего находится на складах":
                data[bc] = int(w.get("quantity", 0)); break
    return data


def _load_ozon_cache() -> dict[str,int]:
    with open(OZON_META_PATH, encoding="utf-8") as f:
        csv_url = json.load(f)["result"]["file"]
    resp = requests.get(csv_url, timeout=30)
    resp.raise_for_status()
    raw = resp.content.decode("utf-8-sig")

    delim = ';' if raw.split("\n",1)[0].count(';') > raw.split("\n",1)[0].count(',') else ','
    reader = csv.DictReader(io.StringIO(raw), delimiter=delim)

    data: dict[str,int] = {}
    for r in reader:
        bc = r.get("Barcode", "").strip()
        if not bc:
            continue
        qty_raw = r.get("Доступно к продаже по схеме FBO, шт.", "0")
        qty = int(float(qty_raw.replace("\xa0","").replace(" ","").replace("\n","") or 0))
        data[bc] = qty
    return data


def _load_wb_supply_cache() -> dict[str,int]:
    """Загружаем данные о поставках из Excel файлов"""
    supply_dir = r"\\194.32.248.34\Shared\Поставки ВБ"
    data: dict[str,int] = {}
    
    for file in glob.glob(os.path.join(supply_dir, "*.xlsx")):
        if "НЕ ТРОГАТЬ.xlsx" in file:
            continue
            
        with open(file, "rb") as f:
            wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            name = str(row[0]).strip() if row[0] else ""
            bc = str(row[1]).strip() if row[1] else ""
            qty = int(row[2] or 0) if row[2] else 0
            
            if bc:
                data[bc] = data.get(bc, 0) + qty
                
        wb.close()
    return data

# ------------- sql‑helpers -------------

_WEEK_BOUNDS_CTE = """
WITH bounds AS (
    SELECT date_trunc('week', CURRENT_DATE)::date  - INTERVAL '7 days'  AS week_start,
           date_trunc('week', CURRENT_DATE)::date  - INTERVAL '1 day'   AS week_end
)
"""

def _load_wb_orders_yesterday(cur):
    cur.execute("""
        SELECT barcode::text, COUNT(*)
        FROM orders_wb
        WHERE date = CURRENT_DATE - INTERVAL '1 day'
        GROUP BY barcode""")
    return {str(bc).split('.')[0]: cnt for bc, cnt in cur.fetchall()}


def _load_wb_orders_week(cur):
    cur.execute(
        _WEEK_BOUNDS_CTE +
        """
        SELECT barcode::text, COUNT(*)
        FROM orders_wb, bounds
        WHERE date BETWEEN bounds.week_start AND bounds.week_end
        GROUP BY barcode"""
    )
    return {str(bc).split('.')[0]: cnt for bc, cnt in cur.fetchall()}


def _load_ozon_orders_week(cur):
    cur.execute(
        _WEEK_BOUNDS_CTE +
        """
        SELECT ref.barcode::text, COALESCE(SUM(oz.ordered_units),0)
        FROM orders_ozon oz
        LEFT JOIN reference ref ON oz.product_sku = ref.ozon_fbo_sku_id, bounds
        WHERE oz.date BETWEEN bounds.week_start AND bounds.week_end
        GROUP BY ref.barcode"""
    )
    return {str(bc).split('.')[0]: cnt for bc, cnt in cur.fetchall() if bc}

# ------------- main refresh -------------

def refresh_data():
    global wh_cache, wb_cache, ozon_cache
    global wb_yes_cache, wb_week_cache, ozon_week_cache, wb_supply_cache, last_refresh
    try:
        with db_conn() as conn, conn.cursor() as cur:
            wb_yes_cache    = _load_wb_orders_yesterday(cur)
            wb_week_cache   = _load_wb_orders_week(cur)
            ozon_week_cache = _load_ozon_orders_week(cur)

        wh_cache  = _load_wh_cache()
        wb_cache  = _load_wb_cache()
        ozon_cache= _load_ozon_cache()
        wb_supply_cache = _load_wb_supply_cache()  # Загружаем данные о поставках

        last_refresh = datetime.now()
        app.logger.info(
            "Refresh OK | wh:%d wb:%d oz:%d | wb_yes:%d wb_week:%d ozon_week:%d | wb_supply:%d",
            len(wh_cache), len(wb_cache), len(ozon_cache),
            len(wb_yes_cache), len(wb_week_cache), len(ozon_week_cache),
            len(wb_supply_cache))
    except Exception as e:
        app.logger.exception("refresh_data failed: %s", e)

# ---------------- scheduler ----------------
sched = BackgroundScheduler(daemon=True)
sched.add_job(refresh_data, 'interval', hours=1, next_run_time=datetime.now())
sched.start()

# --------------- auth decorator ---------------

def require_auth(f):
    def wrapper(*a, **k):
        if request.headers.get('Authorization') != AUTH_TOKEN:
            return abort(401)
        return f(*a, **k)
    wrapper.__name__ = f.__name__; return wrapper

# --------------- endpoints ---------------

@app.route('/barcode')
@require_auth
def barcode_lookup():
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({'error': 'Missing code'}), 400

    wh_entry = wh_cache.get(code, {})
    res = {
        'name'      : wh_entry.get('name'),
        'wh'        : wh_entry.get('stock'),
        'wb'        : wb_cache.get(code),
        'ozon'      : ozon_cache.get(code),
        'wb_yes'    : wb_yes_cache.get(code, 0),
        'wb_week'   : wb_week_cache.get(code, 0),
        'ozon_week' : ozon_week_cache.get(code, 0),
        'wb_supply' : wb_supply_cache.get(code, 0)  # Добавляем данные о поставках
    }
    return jsonify(res)


@app.route('/refresh', methods=['POST'])
@require_auth
def manual_refresh():
    refresh_data()
    return jsonify({'status': 'reloaded', 'time': last_refresh.isoformat()})

# --------------- entry ---------------
if __name__ == '__main__':
    refresh_data()
    app.run(host='0.0.0.0', port=51000, debug=True)
